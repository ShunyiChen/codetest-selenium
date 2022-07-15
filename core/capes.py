import os
import pathlib
import random
import time
import warnings
from tkinter import messagebox

from aip import AipOcr
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common import NoSuchElementException, TimeoutException
from selenium.webdriver import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains

from conf import r_name, w_name
from core import get_now_time


class User:
    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.complete = False


driver = None
aip_ocr = None
users = [User('SADC_01', 'ASDqwe123!'), User('SAC_01', 'ASDqwe123!')]
start_time = get_now_time()
end_time = get_now_time()
_update_progress_label = None


def fill_data(update_status_label, update_progress_label):
    global aip_ocr
    global driver
    global _update_progress_label
    _update_progress_label = update_progress_label
    # 初始化百度OCR服务连接
    aip_ocr = AipOcr("25675193", "aq4ULqZRRzIZHglVpVhbtjDK", "SvKmQjurRTEkKNVqIGlje8TH5kKnp8cl")
    # 加载驱动
    ser = Service(r'C:\Users\QD291NB\Downloads\selenium\103.0.5060.53\chromedriver_win32\chromedriver.exe')
    # 实现不自动关闭的重点
    option = webdriver.ChromeOptions()
    option.add_experimental_option("detach", True)
    driver = webdriver.Chrome(service=ser, options=option)
    # 设置隐式等待时间为10秒
    # driver.implicitly_wait(10)
    loop_login()
    update_status_label("已停止")


def loop_login():
    global start_time
    start_time = get_now_time()
    # 浏览器最大化
    driver.maximize_window()
    # 打开网址
    url = 'http://localhost:8081/#/login'
    driver.get(url)

    i = 0
    for user in users:
        times = 0
        # 允许单用户最大登录次数为10
        while times < 10:
            # 登录系统
            logged_in_user_name = login(user.username, user.password)
            if logged_in_user_name == '' or logged_in_user_name != user.username:
                times += 1
                print('登录失败', times, '次', user.username)
            else:
                # 检索
                retrieves_data_by_year()
                # 导出
                bulk_export()
                # 填写
                fill_and_create_new_template()
                # 导入并提交
                import_and_submit_your_data()
                print("Did " + user.username + " work")
                # 退出系统
                logout(user.username)

                user.complete = True
                break
        i += 1
        _update_progress_label(str(i) + "/" + str(len(users)))
    global end_time
    end_time = get_now_time()
    messagebox.showinfo("提示", "完成!!")
    driver_quit()


def driver_quit():
    if driver is not None:
        driver.quit()


def login(username, password):
    # 定位用户名文本框
    username_field = WebDriverWait(driver, 5, poll_frequency=0.05, ignored_exceptions=NoSuchElementException).until(
        expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, "input[placeholder='账号'][type='text']")))

    # 清除文本框内容
    # username_field.send_keys(Keys.CONTROL + "a")
    # username_field.send_keys(Keys.DELETE)
    driver.execute_script('document.querySelector(".el-input__inner").value=""')
    # 输入用户名
    username_field.send_keys(username)

    # 定位密码文本框
    pass_field = driver.find_element(By.CSS_SELECTOR, "input[placeholder='密码'][type='password']")
    # pass_field = WebDriverWait(driver, 10).until(expected_conditions.visibility_of_element_located((By.CSS_SELECTOR, "input[type='password']")))
    # 清除文本框内容
    pass_field.send_keys(Keys.CONTROL + "a")
    pass_field.send_keys(Keys.DELETE)
    # 输入密码
    pass_field.send_keys(password)

    # 定位验证码图片
    img = driver.find_element(By.XPATH, "//img[@alt='']")
    # 验证码截图
    data = img.screenshot_as_png
    # 通过ORC识别图像
    res = aip_ocr.basicGeneral(data, {})
    # 返回识别结果
    wr = res['words_result']
    code = wr[0]['words']
    validate_code = driver.find_element(By.XPATH, "//input[@placeholder='验证码']")
    # 清除文本框内容
    validate_code.send_keys(Keys.CONTROL + "a")
    validate_code.send_keys(Keys.DELETE)
    # 输入验证码
    validate_code.send_keys(code)

    # 点击登录按钮
    login_button = driver.find_element(By.TAG_NAME, 'button')
    login_button.click()

    try:
        menu = WebDriverWait(driver, 5).until(
            expected_conditions.visibility_of_element_located((By.CSS_SELECTOR,
                                                               "span[role='button'][aria-haspopup='list']")))
        return menu.text
    except TimeoutException:
        pass
    return ''


def logout(username):
    # 点击用户名菜单
    menu = WebDriverWait(driver, 30). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//ul[@class='site-navbar__menu site-navbar__menu--right el-menu--horizontal el-menu']")))
    menu.click()

    logout_item = WebDriverWait(driver, 3). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//li[text()='退出']")))
    logout_item.click()
    ok_button = WebDriverWait(driver, 3). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//button[@class='el-button el-button--default el-button--small el-button--primary ']")))
    ok_button.click()
    print(username, " has logged out")


def retrieves_data_by_year():
    # 点击导航栏"KPI填报"
    kpi_fill_item = WebDriverWait(driver, 3). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//div[@class='el-scrollbar__view']/ul[@role='menubar']/li[2]")))
    kpi_fill_item.click()
    # 点击填报年度下拉框
    fill_year_item = WebDriverWait(driver, 3). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//div[@class='el-form-item__content']/div[@class='el-select el-select--medium']/div[@class='el-input el-input--medium el-input--suffix']")))
    fill_year_item.click()

    # 选择“2023FY”
    select_year_item = WebDriverWait(driver, 3). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//li[1]/span[text()='2023FY']")))
    select_year_item.click()

    # 定位鼠标悬停元素-填报期下拉框
    clear_period_item = driver.find_element(By.XPATH, "//*[@id='pane-fill-fill-kpi']/div/div/div/form/div[1]/div[2]/div/div/div/div/input")
    # 鼠标悬停
    ActionChains(driver).move_to_element(clear_period_item).perform()

    # 点击"填报期下拉框"上关闭图标
    close_icon = WebDriverWait(driver, 3). \
        until(expected_conditions.visibility_of_element_located((By.XPATH, "//*[@id='pane-fill-fill-kpi']/div/div/div/form/div[1]/div[2]/div/div/div/div/span/span")))
    close_icon.click()

    # 点击查询按钮
    driver.find_element(By.XPATH,
                        "//button[@class='el-button el-button--primary el-button--medium']/span[text()='查询']").click()

    WebDriverWait(driver, 5). \
        until(expected_conditions.visibility_of_element_located((By.CLASS_NAME,
                                                                 "el-pagination__total")))


def is_file_downloaded():
    # checks if file downloaded file path exists
    while not os.path.exists(r_name):
        time.sleep(1)
        # check file
        if os.path.isfile(r_name):
            print("File download is completed")
        else:
            print("File download is not completed")


def bulk_export():
    # 删除上次执行过的文件
    r_name_file = pathlib.Path(r_name)
    if r_name_file.exists():
        r_name_file.unlink()
    w_name_file = pathlib.Path(w_name)
    if w_name_file.exists():
        w_name_file.unlink()

    # 点击批量导出按钮
    bulk_export_button = driver.find_element(By.XPATH,
                                             "//button[@class='el-button fillBt el-button--primary el-button--medium']/span[text()='批量导出']")
    bulk_export_button.click()
    # check if file downloaded file path exists
    is_file_downloaded()


def fill_and_create_new_template():
    with warnings.catch_warnings(record=True):
        file_path = os.path.join('./', r_name)
        wb = load_workbook(file_path)
        # sheet = wb.get_sheet_names()
        w1 = wb.get_sheet_by_name('Sheet1')
        rows = w1.max_row
        for i in range(1, rows):
            # 单位列
            unit = w1["H" + str(i + 1)].value
            if unit == '吨':
                w1["G" + str(i + 1)] = random.randint(10000, 100000)
            elif unit == '标准立方米':
                w1["G" + str(i + 1)] = random.randint(10000, 100000)
            elif unit == '起':
                w1["G" + str(i + 1)] = random.randint(10, 252)
            elif unit == '家':
                w1["G" + str(i + 1)] = random.randint(1000, 2520)
            elif unit == '件':
                w1["G" + str(i + 1)] = random.randint(3000, 12520)
            elif unit == '千瓦时':
                w1["G" + str(i + 1)] = random.randint(50000, 1000000)
            elif unit == '万元':
                w1["G" + str(i + 1)] = round(random.uniform(500.01, 1000.99), 2)  # 保留两位小数
            else:
                w1["G" + str(i + 1)] = random.randint(3000, 12520)
        wb.save(w_name)


def import_and_submit_your_data():
    # 导入文件
    element = driver.find_element_by_name('file')
    element.send_keys(w_name)

    # 点击"确定"
    ok_button = WebDriverWait(driver, 30). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//button[@class='el-button el-button--default el-button--small el-button--primary ']")))
    ok_button.click()

    # 点击"批量提交"
    bulk_submit_button = WebDriverWait(driver, 30). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//button[@class='el-button fillBt el-button--primary el-button--medium']/span[text()='批量提交']")))
    bulk_submit_button.click()

    # 点击"确认"
    confirm_button = WebDriverWait(driver, 30). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//button[@class='el-button el-button--default el-button--small el-button--primary ']")))
    confirm_button.click()

    # 先等待“弹出窗口”出现，然后再点击“确定”关闭
    WebDriverWait(driver, 60). \
        until(expected_conditions.visibility_of_element_located((By.XPATH,
                                                                 "//p[text()='批量提交成功']")))
    okay_button = driver.find_element(By.XPATH,
                                      "//button[@class='el-button el-button--default el-button--small el-button--primary ']")
    okay_button.click()


def get_start_time():
    return start_time


def get_end_time():
    return end_time
